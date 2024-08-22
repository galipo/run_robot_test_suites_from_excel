from openpyxl import Workbook,load_workbook

wb=load_workbook('py_deneme.xlsx')
ws=wb.active

print(ws)

j=2
count=0

while count!=4:
    #Take datas
    i = 1
    list=[]
    while i<=5:
        list.append(ws.cell(j,i).value)    #taking data and adding a list from excel file         
        i=i+1
    print(list)                            #print the terminal
    if len(list) == list.count(None):
         count=count+1
    else:
         #Write text
         f = open("py_deneme.txt","a")       #opening for writing the file
           #create a string for fisrt text
         paragh = ('ankara='+str(list[0])+'\n'+'ist='+str(list[1])+'\n'+'fran='+str(list[2])+'\n'+'alm='+str(list[3])+'\n'+'selan='+str(list[4])+'\n')
           #write the text into .txt file
         f.write(paragh)
         f.write('*****************************************************************')
         f.write('\n')
    j=j+1
  
