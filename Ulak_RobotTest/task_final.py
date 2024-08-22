from openpyxl import Workbook,load_workbook

wb = load_workbook("RobotTestSuitesParamsList(deneme).xlsx")
ws = wb.active

#Writing name of the active worksheet        
print(ws)

j=2                                              #first row number
offset=0                                         #finisher variable to loop

while offset!=2:                                 #if offset variable counts two empty(none) row, the loop will end
    #Take datas
    i = 1                                        #first column number
    list=[]                                      #the list for excel datas
    while i<=19:
        list.append(ws.cell(j,i).value)          #taking data and adding a list from excel file         
        i=i+1
    print(list)                                  #print the terminal
    
    if len(list) == list.count(None):            #checking the empty rows
         offset=offset+1
    else:     
         #Write text
         f = open("ArgumentFile.txt","a")        #opening for writing the file

         #create a string for first text
         paragh = '--doc This is a test suite  for Setup:'+str(list[0])+':Mode:'+str(list[1])+':UeType:'+str(list[2])+':UeNum:'+str(list[3])+':Freq:'+str(list[4])+':UeAttachIter:'+str(list[5])+':Dryrun:'+str(list[6])+':DLTxCoeff:'+str(list[7])+':DLRxCoeff:'+str(list[8])+':DuTrV:'+str(list[9])+':CuUpTrV:'+str(list[10])+':subTestSuiteName:'+str(list[11])+':androidUeList:'+str(list[12])+':TagList:'+str(list[13])+':phyRunArg:'+str(list[14])+':ULTxCoeff:'+str(list[15])+':ULRxCoeff:'+str(list[16])+':ConfiguredCellNumber:'+str(list[17])+':ActivatedCellIndexList:'+str(list[18])

         #write the text into .txt file
         f.write(paragh)
         f.write('\n')

         #write other items
         string = str(list[13])
         string = string.split(',')

         f.write('-v Env_Setup_ID:'+str(string[0])+'\n')
         f.write('-v Env_mode:'+str(list[1])+'\n')
         f.write('-v Env_ue_type:'+str(list[2])+'\n')
         f.write('-v Env_ue_number:'+str(list[3])+'\n')
         f.write('-v Env_frequency:'+str(list[4])+'\n')
         f.write('-v Env_UeAttachIter:'+str(list[5])+'\n')
         f.write('-v Env_iperf_dl_tx_coeff:'+str(list[7])+'\n')
         f.write('-v Env_iperf_dl_rx_coeff:'+str(list[8])+'\n')
         f.write('-v Env_du_tr_verbosity:'+str(list[9])+'\n')
         f.write('-v Env_subTestSuiteName:'+str(list[11])+'\n')
         f.write('-v Env_androidUeList:'+str(list[12])+'\n')
         
         k=0
         while k<=(len(string)-1):
           f.write('-i '+str(string[k])+'\n')
           k=k+1
        
         f.write('-v Env_pull_request_flag:'+str(list[6])+'\n')
         f.write('-v Env_phyRunArg:'+str(list[14])+'\n')
         f.write('-v Env_iperf_ul_tx_coeff:'+str(list[15])+'\n')
         f.write('-v Env_iperf_ul_rx_coeff:'+str(list[16])+'\n')
         f.write('-v Env_configured_cell_number:'+str(list[17])+'\n')
         f.write('-v Env_activated_cell_index_list:'+str(list[18])+'\n')

         f.write('--name '+str(string[0])+str(list[1])+str(list[11])+str(list[3])+str(list[2])+str(list[1])+'Ue'+str(list[4])+'\n')
         f.write('-l '+str(string[0])+str(list[1])+str(list[11])+str(list[3])+str(list[2])+str(list[1])+'Ue'+str(list[4])+'_log.html'+'\n')
         f.write('-o '+str(string[0])+str(list[1])+str(list[11])+str(list[3])+str(list[2])+str(list[1])+'Ue'+str(list[4])+'_output.xml'+'\n')
         f.write('-r '+str(string[0])+str(list[1])+str(list[11])+str(list[3])+str(list[2])+str(list[1])+'Ue'+str(list[4])+'_report.html'+'\n')

         f.write('-v Env_reboot:'+str(list[6])+'\n')

         f.write('--exitonfailure'+'\n')

         f.write('-v Env_deploy_directory:/home/ulak/ulak-gnb/fx_2207/bin/nr5g/gnb/DAILY_BUILDS/develop'+'\n')
         f.write('-v Env_cu_run_script:cu_run_robot_develop.sh'+'\n')
         f.write('-v Env_du_run_script:l2_run_robot_develop.sh'+'\n')
         f.write('-v Env_oamc_run_script:oamc_run_robot_develop.sh'+'\n')
         f.write('*****************************************************************'+'\n')
         f.close()
    j=j+1